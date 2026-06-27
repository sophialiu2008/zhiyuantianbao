#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Clean the 2026 undergraduate physics admission-plan workbook with precomputed
2023-2025 score/rank history into app-ready JSON/CSV files.

Usage:
  python scripts/data_cleaning/clean_admission_plan_rank_xlsx.py \
    --input "C:/Users/liuli/Desktop/26年本科批物理组_招生计划_含历年位次.xlsx" \
    --output data/cleaned/admission_plan
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PLAN_SHEET_NAME = "26年招生计划_含历年位次"
DEFAULT_BATCH = "本科批"
DEFAULT_PLAN_NATURE = "非定向"
DEFAULT_SUBJECT = "physics"
DEFAULT_VOLUNTEER_MODE = "平行志愿"
HISTORY_YEARS = (2025, 2024, 2023)


@dataclass
class AdmissionPlan:
    year: int
    subject: str
    batch: str
    plan_nature: str
    volunteer_mode: str
    school_code: str
    school_name: str
    school_name_raw: str
    school_name_normalized: str
    school_tags: list[str]
    major_code: str
    major_name: str
    major_name_normalized: str
    major_remark: str
    subject_requirement: str
    plan_count: int | None
    duration_years: int | None
    tuition: int | None
    source_file: str
    source_page: int | None
    source_row: int | None
    entry: str
    crawled_at: str | None
    raw_payload: dict[str, Any]


@dataclass
class AdmissionPlanHistoryMatch:
    plan_year: int
    subject: str
    batch: str
    school_code: str
    school_name: str
    school_name_raw: str
    school_name_normalized: str
    major_code: str
    major_name: str
    major_name_normalized: str
    major_remark: str
    min_score_2025: int | None
    min_rank_2025: int | None
    min_score_2024: int | None
    min_rank_2024: int | None
    min_score_2023: int | None
    min_rank_2023: int | None
    history_years: int
    best_rank: int | None
    worst_rank: int | None
    avg_rank: int | None
    latest_score: int | None
    latest_rank: int | None
    is_new_major: bool
    match_note: str
    source_file: str
    source_row: int
    raw_payload: dict[str, Any]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("_x000D_", "").replace("\ufeff", "").strip()
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def to_int(value: Any) -> int | None:
    text = clean_text(value)
    if not text:
        return None
    match = re.search(r"-?\d+(?:\.0+)?", text)
    if not match:
        return None
    try:
        return int(float(match.group(0)))
    except ValueError:
        return None


def normalize_code(value: Any, width: int | None = None) -> str:
    text = clean_text(value)
    if re.fullmatch(r"\d+(?:\.0+)?", text):
        text = str(int(float(text)))
    if width and text.isdigit():
        return text.zfill(width)
    return text


def normalize_school_name(value: str) -> str:
    text = re.sub(r"\[[^\]]*\]", "", value)
    text = re.sub(
        r"\(([^)]*(市|地方专项|国家专项|中外合作|按高考文化成绩|少数民族预科|国际合作|国际本科|职教)[^)]*)\)",
        "",
        text,
    )
    text = re.sub(r"\s+", "", text)
    return text.strip()


def normalize_major_name(value: str) -> str:
    text = re.sub(r"[（(].*?[）)]", "", value)
    text = re.sub(r"\s+", "", text)
    return text.strip()


def split_school_name(raw: str) -> tuple[str, list[str]]:
    tags = re.findall(r"\[([^\]]+)\]", raw)
    name = re.sub(r"\[[^\]]*\]", "", raw).strip()
    return name, tags


def row_payload(headers: list[str], values: list[Any]) -> dict[str, str]:
    return {clean_text(header): clean_text(value) for header, value in zip(headers, values) if clean_text(header)}


def history_summary(values: dict[int, tuple[int | None, int | None]]) -> dict[str, int | None]:
    ranks = [rank for _, rank in values.values() if rank is not None]
    if not ranks:
        return {
            "history_years": 0,
            "best_rank": None,
            "worst_rank": None,
            "avg_rank": None,
            "latest_score": None,
            "latest_rank": None,
        }
    latest_year = next(year for year in HISTORY_YEARS if values[year][1] is not None)
    return {
        "history_years": len(ranks),
        "best_rank": min(ranks),
        "worst_rank": max(ranks),
        "avg_rank": round(sum(ranks) / len(ranks)),
        "latest_score": values[latest_year][0],
        "latest_rank": values[latest_year][1],
    }


def parse_workbook(path: Path, year: int, sheet_name: str) -> tuple[list[AdmissionPlan], list[AdmissionPlanHistoryMatch]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise SystemExit(f"Sheet not found: {sheet_name}. Available sheets: {', '.join(workbook.sheetnames)}")
    sheet = workbook[sheet_name]

    headers = [clean_text(cell.value) for cell in next(sheet.iter_rows(min_row=3, max_row=3))]
    plans: list[AdmissionPlan] = []
    matches: list[AdmissionPlanHistoryMatch] = []

    for excel_row, cells in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
        values = list(cells[: len(headers)])
        school_code = normalize_code(values[1], width=4)
        school_name_raw = clean_text(values[2])
        major_code = normalize_code(values[3], width=2)
        major_name = clean_text(values[4])
        if not school_code or not school_name_raw or not major_code or not major_name:
            continue

        school_name, school_tags = split_school_name(school_name_raw)
        major_remark = clean_text(values[5])
        subject_requirement = clean_text(values[6]) or "不限"
        payload = row_payload(headers, values)
        source_row = to_int(values[0]) or excel_row

        plan = AdmissionPlan(
            year=year,
            subject=DEFAULT_SUBJECT,
            batch=DEFAULT_BATCH,
            plan_nature=DEFAULT_PLAN_NATURE,
            volunteer_mode=DEFAULT_VOLUNTEER_MODE,
            school_code=school_code,
            school_name=school_name,
            school_name_raw=school_name_raw,
            school_name_normalized=normalize_school_name(school_name_raw),
            school_tags=school_tags,
            major_code=major_code,
            major_name=major_name,
            major_name_normalized=normalize_major_name(major_name),
            major_remark=major_remark,
            subject_requirement=subject_requirement,
            plan_count=to_int(values[7]),
            duration_years=to_int(values[8]),
            tuition=to_int(values[9]),
            source_file=path.name,
            source_page=None,
            source_row=source_row,
            entry="rank_xlsx",
            crawled_at=None,
            raw_payload=payload,
        )
        plans.append(plan)

        history_values = {
            2025: (to_int(values[10]), to_int(values[11])),
            2024: (to_int(values[12]), to_int(values[13])),
            2023: (to_int(values[14]), to_int(values[15])),
        }
        summary = history_summary(history_values)
        note = clean_text(values[16])
        matches.append(
            AdmissionPlanHistoryMatch(
                plan_year=year,
                subject=DEFAULT_SUBJECT,
                batch=DEFAULT_BATCH,
                school_code=school_code,
                school_name=school_name,
                school_name_raw=school_name_raw,
                school_name_normalized=normalize_school_name(school_name_raw),
                major_code=major_code,
                major_name=major_name,
                major_name_normalized=normalize_major_name(major_name),
                major_remark=major_remark,
                min_score_2025=history_values[2025][0],
                min_rank_2025=history_values[2025][1],
                min_score_2024=history_values[2024][0],
                min_rank_2024=history_values[2024][1],
                min_score_2023=history_values[2023][0],
                min_rank_2023=history_values[2023][1],
                history_years=summary["history_years"] or 0,
                best_rank=summary["best_rank"],
                worst_rank=summary["worst_rank"],
                avg_rank=summary["avg_rank"],
                latest_score=summary["latest_score"],
                latest_rank=summary["latest_rank"],
                is_new_major="新增" in note,
                match_note=note,
                source_file=path.name,
                source_row=source_row,
                raw_payload=payload,
            )
        )

    return plans, matches


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    fieldnames = list(rows[0].keys())
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            row = dict(row)
            for key, value in list(row.items()):
                if isinstance(value, (dict, list)):
                    row[key] = json.dumps(value, ensure_ascii=False)
            writer.writerow(row)


def build_report(input_path: Path, output_dir: Path, plans: list[AdmissionPlan], matches: list[AdmissionPlanHistoryMatch]) -> dict[str, Any]:
    subject_requirements = Counter(plan.subject_requirement for plan in plans)
    return {
        "input_file": str(input_path),
        "output_dir": str(output_dir),
        "plan_rows": len(plans),
        "history_match_rows": len(matches),
        "plan_count_total": sum(plan.plan_count or 0 for plan in plans),
        "new_major_count": sum(1 for match in matches if match.is_new_major),
        "any_history_count": sum(1 for match in matches if match.history_years > 0),
        "all_3_year_history_count": sum(1 for match in matches if match.history_years == 3),
        "missing_plan_count": sum(1 for plan in plans if plan.plan_count is None),
        "subject_requirement_distribution": dict(subject_requirements.most_common()),
    }


def write_markdown_report(path: Path, report: dict[str, Any]) -> None:
    lines = [
        "# 2026 本科批物理组招生计划历史位次清洗报告",
        "",
        f"- 输入文件：`{report['input_file']}`",
        f"- 输出目录：`{report['output_dir']}`",
        f"- 计划行数：{report['plan_rows']}",
        f"- 历史快照行数：{report['history_match_rows']}",
        f"- 计划总数：{report['plan_count_total']}",
        f"- 新增专业：{report['new_major_count']}",
        f"- 任一年有历史位次：{report['any_history_count']}",
        f"- 三年都有历史位次：{report['all_3_year_history_count']}",
        f"- 计划数缺失：{report['missing_plan_count']}",
        "",
        "## 再选科目分布",
        "",
    ]
    for subject_requirement, count in report["subject_requirement_distribution"].items():
        lines.append(f"- {subject_requirement}: {count}")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="Workbook path")
    parser.add_argument("--output", default="data/cleaned/admission_plan", help="Output directory")
    parser.add_argument("--year", type=int, default=2026)
    parser.add_argument("--sheet", default=PLAN_SHEET_NAME)
    args = parser.parse_args()

    input_path = Path(args.input).resolve()
    output_dir = Path(args.output).resolve()
    plans, matches = parse_workbook(input_path, args.year, args.sheet)
    plan_rows = [asdict(plan) for plan in plans]
    match_rows = [asdict(match) for match in matches]

    write_json(output_dir / f"admission_plan_{args.year}_physics_benke.json", plan_rows)
    write_csv(output_dir / f"admission_plan_{args.year}_physics_benke.csv", plan_rows)
    write_json(output_dir / f"admission_plan_history_matches_{args.year}_physics_benke.json", match_rows)
    write_csv(output_dir / f"admission_plan_history_matches_{args.year}_physics_benke.csv", match_rows)

    report = build_report(input_path, output_dir, plans, matches)
    write_json(output_dir / "reports" / f"admission_plan_history_matches_{args.year}_physics_benke_report.json", report)
    write_markdown_report(output_dir / "reports" / f"admission_plan_history_matches_{args.year}_physics_benke_report.md", report)

    print(f"Plan rows: {len(plans)}")
    print(f"History match rows: {len(matches)}")
    print(f"Report: {output_dir / 'reports' / f'admission_plan_history_matches_{args.year}_physics_benke_report.md'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
