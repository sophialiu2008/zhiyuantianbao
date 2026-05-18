#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Clean Hebei gaokao admission/rank Excel files into app-ready JSON.

Run from the project data directory:
  python scripts/data_cleaning/clean_excel.py --input . --output data/cleaned
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Iterable

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "Missing dependency: openpyxl. Install it with `pip install openpyxl`."
    ) from exc


SUBJECT_PHYSICS = "physics"
SUBJECT_HISTORY = "history"
SUBJECT_LABELS = {
    SUBJECT_PHYSICS: "物理",
    SUBJECT_HISTORY: "历史",
}


@dataclass
class RankRecord:
    year: int
    subject: str
    score: int
    count_at_score: int
    cumulative_rank: int
    score_label: str
    source_file: str
    source_row: int


@dataclass
class AdmissionRecord:
    year: int
    subject: str
    admission_type: str
    school_code: str
    school_name: str
    school_name_raw: str
    school_tags: list[str]
    major_code: str
    major_name: str
    min_score: int | None
    min_rank: int | None
    chinese_math_score: int | None
    chinese_math_highest: int | None
    foreign_language_score: int | None
    first_choice_subject_score: int | None
    second_choice_subject_highest: int | None
    second_choice_subject_second: int | None
    volunteer_no: int | None
    remark: str
    source_file: str
    source_row: int


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = text.replace("_x000D_", "")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def to_int(value: Any) -> int | None:
    text = clean_text(value)
    if not text:
        return None
    match = re.search(r"-?\d+(?:\.0+)?", text)
    if not match:
        return None
    number = match.group(0)
    try:
        return int(float(number))
    except ValueError:
        return None


def parse_year_from_text(text: str) -> int | None:
    match = re.search(r"(20\d{2})", text)
    if match:
        return int(match.group(1))
    match = re.search(r"(^|[^\d])(\d{2})年", text)
    if match:
        return 2000 + int(match.group(2))
    return None


def infer_year(path: Path, worksheet: Any) -> int:
    year = parse_year_from_text(path.name)
    if year:
        return year
    for row in worksheet.iter_rows(min_row=1, max_row=min(5, worksheet.max_row), values_only=True):
        year = parse_year_from_text(" ".join(clean_text(v) for v in row))
        if year:
            return year
    raise ValueError(f"Cannot infer year from {path.name}")


def classify_workbook(path: Path) -> str | None:
    if path.name.startswith("~$"):
        return None
    if "第一志愿录取后" in path.name:
        return "ignored_rank_after_first_choice"
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None
    try:
        ws = wb.active
        first_rows = []
        for row in ws.iter_rows(min_row=1, max_row=min(8, ws.max_row), values_only=True):
            first_rows.append(" ".join(clean_text(v) for v in row))
        text = "\n".join(first_rows)
        if "分数档次" in text and "累计人数" in text:
            return "rank"
        if "投档" in text and "院校" in text and "专业" in text:
            return "admission"
        return None
    finally:
        wb.close()


def parse_school_name(raw: str) -> tuple[str, list[str]]:
    tags = re.findall(r"\[([^\]]+)\]", raw)
    name = re.sub(r"\[[^\]]+\]", "", raw).strip()
    return name, tags


def parse_rank_workbook(path: Path) -> list[RankRecord]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        year = infer_year(path, ws)
        records: list[RankRecord] = []
        for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_index <= 2:
                continue
            score_label = clean_text(row[0] if len(row) > 0 else "")
            score = to_int(score_label)
            if score is None:
                continue
            physics_count = to_int(row[1] if len(row) > 1 else None)
            physics_rank = to_int(row[2] if len(row) > 2 else None)
            history_count = to_int(row[3] if len(row) > 3 else None)
            history_rank = to_int(row[4] if len(row) > 4 else None)
            if physics_count is not None and physics_rank is not None:
                records.append(
                    RankRecord(
                        year=year,
                        subject=SUBJECT_PHYSICS,
                        score=score,
                        count_at_score=physics_count,
                        cumulative_rank=physics_rank,
                        score_label=score_label,
                        source_file=path.name,
                        source_row=row_index,
                    )
                )
            if history_count is not None and history_rank is not None:
                records.append(
                    RankRecord(
                        year=year,
                        subject=SUBJECT_HISTORY,
                        score=score,
                        count_at_score=history_count,
                        cumulative_rank=history_rank,
                        score_label=score_label,
                        source_file=path.name,
                        source_row=row_index,
                    )
                )
        return records
    finally:
        wb.close()


def build_rank_lookup(records: Iterable[RankRecord]) -> dict[tuple[int, str, int], int]:
    return {(r.year, r.subject, r.score): r.cumulative_rank for r in records}


def parse_admission_workbook(
    path: Path,
    rank_lookup: dict[tuple[int, str, int], int],
) -> list[AdmissionRecord]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        year = infer_year(path, ws)
        current_subject: str | None = None
        current_admission_type = ""
        records: list[AdmissionRecord] = []

        for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
            values = [clean_text(v) for v in row]
            joined = " ".join(values)

            if "物理科目组合" in joined:
                current_subject = SUBJECT_PHYSICS
            elif "历史科目组合" in joined:
                current_subject = SUBJECT_HISTORY

            if values and values[0] in {"非定向", "定向"}:
                current_admission_type = values[0]

            school_code = values[0] if len(values) > 0 else ""
            school_name_raw = values[1] if len(values) > 1 else ""
            major_code = values[2] if len(values) > 2 else ""
            major_name = values[3] if len(values) > 3 else ""

            if not re.fullmatch(r"\d{4}", school_code):
                continue
            if not school_name_raw or not major_code or not major_name:
                continue

            subject = current_subject or SUBJECT_PHYSICS
            school_name, school_tags = parse_school_name(school_name_raw)
            min_score = to_int(values[4] if len(values) > 4 else None)
            min_rank = rank_lookup.get((year, subject, min_score)) if min_score is not None else None

            records.append(
                AdmissionRecord(
                    year=year,
                    subject=subject,
                    admission_type=current_admission_type or "非定向",
                    school_code=school_code,
                    school_name=school_name,
                    school_name_raw=school_name_raw,
                    school_tags=school_tags,
                    major_code=major_code,
                    major_name=major_name,
                    min_score=min_score,
                    min_rank=min_rank,
                    chinese_math_score=to_int(values[5] if len(values) > 5 else None),
                    chinese_math_highest=to_int(values[6] if len(values) > 6 else None),
                    foreign_language_score=to_int(values[7] if len(values) > 7 else None),
                    first_choice_subject_score=to_int(values[8] if len(values) > 8 else None),
                    second_choice_subject_highest=to_int(values[9] if len(values) > 9 else None),
                    second_choice_subject_second=to_int(values[10] if len(values) > 10 else None),
                    volunteer_no=to_int(values[11] if len(values) > 11 else None),
                    remark=values[12] if len(values) > 12 else "",
                    source_file=path.name,
                    source_row=row_index,
                )
            )
        return records
    finally:
        wb.close()


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def write_csv(path: Path, records: list[Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not records:
        path.write_text("", encoding="utf-8")
        return
    rows = [asdict(record) for record in records]
    fieldnames = list(rows[0].keys())
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def report_for_records(
    input_dir: Path,
    output_dir: Path,
    rank_files: list[Path],
    admission_files: list[Path],
    ignored_files: list[Path],
    rank_records: list[RankRecord],
    admission_records: list[AdmissionRecord],
) -> dict[str, Any]:
    rank_by_year_subject = Counter((r.year, r.subject) for r in rank_records)
    admission_by_year_subject = Counter((r.year, r.subject) for r in admission_records)
    missing_score = [r for r in admission_records if r.min_score is None]
    missing_rank = [r for r in admission_records if r.min_score is not None and r.min_rank is None]
    duplicate_keys = Counter(
        (r.year, r.subject, r.school_code, r.major_code, r.major_name, r.admission_type)
        for r in admission_records
    )
    duplicate_count = sum(1 for count in duplicate_keys.values() if count > 1)

    score_ranges: dict[str, dict[str, int | None]] = {}
    for year in sorted({r.year for r in admission_records}):
        year_scores = [r.min_score for r in admission_records if r.year == year and r.min_score is not None]
        score_ranges[str(year)] = {
            "min": min(year_scores) if year_scores else None,
            "max": max(year_scores) if year_scores else None,
        }

    examples = []
    for record in admission_records:
        if (
            record.year == 2025
            and record.subject == SUBJECT_PHYSICS
            and record.school_name == "北京林业大学"
            and record.major_name == "人工智能"
        ):
            examples.append(asdict(record))

    return {
        "input_dir": str(input_dir),
        "output_dir": str(output_dir),
        "rank_files": [p.name for p in rank_files],
        "admission_files": [p.name for p in admission_files],
        "ignored_files": [p.name for p in ignored_files],
        "rank_record_count": len(rank_records),
        "admission_record_count": len(admission_records),
        "rank_by_year_subject": {
            f"{year}_{subject}": count
            for (year, subject), count in sorted(rank_by_year_subject.items())
        },
        "admission_by_year_subject": {
            f"{year}_{subject}": count
            for (year, subject), count in sorted(admission_by_year_subject.items())
        },
        "admission_score_ranges": score_ranges,
        "missing_min_score_count": len(missing_score),
        "missing_min_rank_count": len(missing_rank),
        "duplicate_key_count": duplicate_count,
        "missing_rank_examples": [asdict(r) for r in missing_rank[:20]],
        "known_example_2025_bjfu_ai": examples,
    }


def write_markdown_report(path: Path, report: dict[str, Any]) -> None:
    lines: list[str] = []
    lines.append("# 河北高考数据清洗报告")
    lines.append("")
    lines.append(f"- 输入目录：`{report['input_dir']}`")
    lines.append(f"- 输出目录：`{report['output_dir']}`")
    lines.append(f"- 位次表文件：{', '.join(report['rank_files'])}")
    lines.append(f"- 投档表文件：{', '.join(report['admission_files'])}")
    if report["ignored_files"]:
        lines.append(f"- 已忽略文件：{', '.join(report['ignored_files'])}")
    lines.append("")
    lines.append("## 总量")
    lines.append("")
    lines.append(f"- 位次记录：{report['rank_record_count']}")
    lines.append(f"- 投档记录：{report['admission_record_count']}")
    lines.append(f"- 投档最低分缺失：{report['missing_min_score_count']}")
    lines.append(f"- 投档最低分有值但未匹配到位次：{report['missing_min_rank_count']}")
    lines.append(f"- 重复键数量：{report['duplicate_key_count']}")
    lines.append("")
    lines.append("## 位次表分布")
    lines.append("")
    for key, count in report["rank_by_year_subject"].items():
        lines.append(f"- {key}: {count}")
    lines.append("")
    lines.append("## 投档表分布")
    lines.append("")
    for key, count in report["admission_by_year_subject"].items():
        lines.append(f"- {key}: {count}")
    lines.append("")
    lines.append("## 投档分范围")
    lines.append("")
    for year, value in report["admission_score_ranges"].items():
        lines.append(f"- {year}: {value['min']} - {value['max']}")
    lines.append("")
    lines.append("## 样例校验")
    lines.append("")
    examples = report["known_example_2025_bjfu_ai"]
    if examples:
        item = examples[0]
        lines.append(
            "- 2025 物理 北京林业大学 人工智能："
            f"投档最低分 {item['min_score']}，位次 {item['min_rank']}，"
            f"专业代号 {item['major_code']}，计划/志愿号字段 {item['volunteer_no']}。"
        )
    else:
        lines.append("- 未找到 2025 物理 北京林业大学 人工智能。")
    lines.append("")
    if report["missing_rank_examples"]:
        lines.append("## 未匹配位次样例")
        lines.append("")
        for item in report["missing_rank_examples"]:
            lines.append(
                f"- {item['year']} {SUBJECT_LABELS.get(item['subject'], item['subject'])} "
                f"{item['school_name']} {item['major_name']} 分数 {item['min_score']} "
                f"来源行 {item['source_row']}"
            )
        lines.append("")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default=".", help="Directory containing source Excel files")
    parser.add_argument("--output", default="data/cleaned", help="Directory for cleaned output")
    args = parser.parse_args()

    input_dir = Path(args.input).resolve()
    output_dir = Path(args.output).resolve()

    excel_files = sorted(input_dir.glob("*.xlsx"), key=lambda p: p.name)
    classified: dict[str, list[Path]] = defaultdict(list)
    for path in excel_files:
        kind = classify_workbook(path)
        if kind:
            classified[kind].append(path)

    rank_files = classified["rank"]
    admission_files = classified["admission"]
    if not rank_files:
        raise SystemExit("No rank Excel files found.")
    if not admission_files:
        raise SystemExit("No admission Excel files found.")

    rank_records: list[RankRecord] = []
    for path in rank_files:
        rank_records.extend(parse_rank_workbook(path))
    rank_lookup = build_rank_lookup(rank_records)

    admission_records: list[AdmissionRecord] = []
    for path in admission_files:
        admission_records.extend(parse_admission_workbook(path, rank_lookup))

    for year in sorted({r.year for r in rank_records}):
        year_records = [r for r in rank_records if r.year == year]
        write_json(output_dir / "rank_table" / f"rank_{year}.json", [asdict(r) for r in year_records])
        write_csv(output_dir / "rank_table" / f"rank_{year}.csv", year_records)

    for year in sorted({r.year for r in admission_records}):
        year_records = [r for r in admission_records if r.year == year]
        write_json(output_dir / "admission" / f"admission_{year}.json", [asdict(r) for r in year_records])
        write_csv(output_dir / "admission" / f"admission_{year}.csv", year_records)

    write_json(output_dir / "rank_table" / "all_rank.json", [asdict(r) for r in rank_records])
    write_json(output_dir / "admission" / "all_admission.json", [asdict(r) for r in admission_records])

    report = report_for_records(
        input_dir=input_dir,
        output_dir=output_dir,
        rank_files=rank_files,
        admission_files=admission_files,
        ignored_files=classified.get("ignored_rank_after_first_choice", []),
        rank_records=rank_records,
        admission_records=admission_records,
    )
    write_json(output_dir / "reports" / "data_cleaning_report.json", report)
    write_markdown_report(output_dir / "reports" / "data_cleaning_report.md", report)

    print(f"Rank files: {len(rank_files)}")
    print(f"Admission files: {len(admission_files)}")
    print(f"Ignored files: {len(classified.get('ignored_rank_after_first_choice', []))}")
    print(f"Rank records: {len(rank_records)}")
    print(f"Admission records: {len(admission_records)}")
    print(f"Report: {output_dir / 'reports' / 'data_cleaning_report.md'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
