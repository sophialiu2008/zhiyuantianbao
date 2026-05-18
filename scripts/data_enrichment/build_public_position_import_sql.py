from __future__ import annotations

import csv
import json
import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
OUT_DIR = ROOT / "data" / "enriched" / "public_positions"
CHUNK_DIR = OUT_DIR / "import_chunks"
MAJORS_JSON = OUT_DIR / "major_profiles_for_matching.json"
CIVIL_XLSX = ROOT / "河北省公务员考试.xlsx"
MILITARY_XLSX = ROOT / "军队文职招考.xlsx"

SUMMARY_JSON = OUT_DIR / "public_position_import_summary.json"
MATCH_REVIEW_CSV = OUT_DIR / "public_position_match_summary.csv"

SQL_CHUNK_SIZE = 600
MATCH_CHUNK_SIZE = 3500


@dataclass(frozen=True)
class Major:
    major_code: str | None
    major_name: str
    normalized_name: str
    keywords: tuple[str, ...]
    categories: tuple[str, ...]


def clean_text(value: Any) -> str | None:
    if value is None or pd.isna(value):
        return None
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "null", "--"}:
        return None
    return text


def clean_int(value: Any) -> int | None:
    if value is None or pd.isna(value):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def normalize_major_name(name: str | None) -> str:
    text = clean_text(name) or ""
    text = re.sub(r"[（(].*?[）)]", "", text)
    return re.sub(r"\s+", "", text)


def compact_text(text: str | None) -> str:
    return re.sub(r"\s+", "", clean_text(text) or "")


def sql_literal(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, int):
        return str(value)
    text = (
        str(value)
        .replace("\\", "\\\\")
        .replace("'", "''")
        .replace("\r", "\\r")
        .replace("\n", "\\n")
        .replace("\t", "\\t")
    )
    return f"E'{text}'"


def json_sql(value: dict[str, Any]) -> str:
    return f"{sql_literal(json.dumps(value, ensure_ascii=False, default=str))}::jsonb"


RULES: list[tuple[tuple[str, ...], tuple[str, ...], tuple[str, ...]]] = [
    (("计算机", "软件", "网络工程", "信息安全", "网络空间安全", "数据科学", "大数据", "智能科学"), ("计算机类", "计算机科学与技术", "软件工程", "网络工程", "信息安全", "网络空间安全", "大数据技术与工程"), ("计算机", "软件", "网络", "信息安全", "大数据")),
    (("人工智能", "智能"), ("计算机类", "电子信息类", "自动化类", "人工智能", "智能科学与技术"), ("人工智能", "智能", "算法", "计算机")),
    (("电子", "通信", "信息工程", "微电子", "集成电路"), ("电子信息类", "电子科学与技术", "信息与通信工程", "通信工程", "集成电路"), ("电子", "通信", "信号", "集成电路", "微电子")),
    (("自动化", "控制"), ("自动化类", "控制科学与工程", "控制工程"), ("自动化", "控制", "机器人")),
    (("电气", "电力"), ("电气类", "电气工程", "电气工程及其自动化"), ("电气", "电力", "电网")),
    (("机械", "车辆", "汽车", "智能制造"), ("机械类", "机械工程", "机械设计制造及其自动化", "车辆工程"), ("机械", "车辆", "制造")),
    (("土木", "建筑", "工程管理", "工程造价"), ("土木类", "土木工程", "建筑类", "管理科学与工程类"), ("土木", "建筑", "工程管理", "工程造价")),
    (("材料", "高分子"), ("材料类", "材料科学与工程", "材料工程"), ("材料", "高分子")),
    (("化学", "化工", "制药"), ("化学类", "化工与制药类", "化学工程与技术"), ("化学", "化工", "制药")),
    (("数学", "信息与计算科学"), ("数学类", "数学", "应用数学"), ("数学", "数理")),
    (("统计", "应用统计"), ("统计学类", "统计学", "应用统计"), ("统计", "数据分析")),
    (("会计", "财务管理", "审计"), ("会计学", "财务管理", "审计学", "工商管理类", "会计"), ("会计", "财务", "审计")),
    (("金融", "保险", "投资"), ("金融学类", "金融学", "金融", "应用经济学"), ("金融", "保险", "证券", "投资")),
    (("经济", "财政", "税收", "国际经济"), ("经济学类", "财政学类", "经济学", "理论经济学", "应用经济学"), ("经济", "财政", "税收")),
    (("法学", "知识产权", "法律"), ("法学类", "法律类", "法律硕士", "法学"), ("法学", "法律", "司法")),
    (("公安", "侦查", "治安"), ("公安学类", "公安技术类", "侦查学"), ("公安", "侦查", "治安")),
    (("汉语言", "中文", "秘书", "新闻", "传播"), ("中国语言文学类", "新闻传播学类", "汉语言文学", "新闻学"), ("中文", "文字", "宣传", "新闻")),
    (("英语", "翻译", "商务英语"), ("外国语言文学类", "英语", "翻译"), ("英语", "翻译", "外语")),
    (("临床医学", "医学影像", "麻醉", "口腔医学"), ("临床医学类", "临床医学", "口腔医学类", "医学影像学", "麻醉学"), ("临床", "医学", "医师")),
    (("护理",), ("护理学", "护理类", "护理"), ("护理", "护士")),
    (("药学", "中药"), ("药学类", "中药学类", "药学"), ("药学", "药剂")),
    (("公共管理", "行政管理", "劳动与社会保障"), ("公共管理类", "行政管理", "公共事业管理"), ("行政", "公共管理")),
    (("工商管理", "市场营销", "人力资源"), ("工商管理类", "工商管理", "市场营销", "人力资源管理"), ("管理", "市场", "人力资源")),
    (("教育", "心理"), ("教育学类", "心理学类", "教育学", "心理学"), ("教育", "心理")),
    (("马克思", "思想政治"), ("马克思主义理论类", "政治学类", "马克思主义理论"), ("政治", "党务", "思想政治")),
]


def infer_major_terms(name: str) -> tuple[tuple[str, ...], tuple[str, ...]]:
    normalized = normalize_major_name(name)
    categories: set[str] = set()
    keywords: set[str] = set()
    if normalized:
        keywords.add(normalized)
    for triggers, cats, keys in RULES:
        if any(trigger in normalized for trigger in triggers):
            categories.update(cats)
            keywords.update(keys)
    if normalized.endswith("学"):
        keywords.add(normalized[:-1])
    return tuple(sorted(keywords, key=len, reverse=True)), tuple(sorted(categories, key=len, reverse=True))


def load_majors() -> list[Major]:
    rows = json.loads(MAJORS_JSON.read_text(encoding="utf-8"))
    majors: list[Major] = []
    seen: set[str] = set()
    for row in rows:
        name = clean_text(row.get("major_name"))
        if not name:
            continue
        normalized = normalize_major_name(name)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        keywords, categories = infer_major_terms(name)
        majors.append(Major(row.get("major_code"), name, normalized, keywords, categories))
    return majors


def read_civil_positions() -> list[dict[str, Any]]:
    df = pd.read_excel(CIVIL_XLSX, header=1)
    df = df.dropna(how="all")
    rows: list[dict[str, Any]] = []
    for _, record in df.iterrows():
        position_code = clean_text(record.get("职位代码"))
        if not position_code:
            continue
        rows.append({
            "position_code": position_code,
            "exam_area": clean_text(record.get("考区")),
            "department": clean_text(record.get("部门")),
            "unit_name": clean_text(record.get("单位")),
            "position_name": clean_text(record.get("职位名称")),
            "recruit_count": clean_int(record.get("招录人数")),
            "education_min": clean_text(record.get("学历低限")),
            "degree_min": clean_text(record.get("学位低限")),
            "major_requirement": clean_text(record.get("专业要求")),
            "other_requirement": clean_text(record.get("其他要求")),
            "essay_type": clean_text(record.get("申论试卷类型")),
            "raw_payload": {k: (None if pd.isna(v) else v) for k, v in record.to_dict().items()},
        })
    return rows


def read_military_positions() -> list[dict[str, Any]]:
    wb = load_workbook(MILITARY_XLSX, read_only=True, data_only=True)
    ws = wb.active
    cols = [
        "position_code", "employer_no", "employer_name", "position_category", "position_name",
        "work_content", "recruit_count", "shortlist_ratio", "source_category", "education",
        "degree", "major_requirement", "exam_subject", "title_graduate", "title_social",
        "qualification_graduate", "qualification_social", "other_requirement", "work_location", "contact_phone",
    ]
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=7, max_col=20, values_only=True):
        if not any(v is not None for v in values):
            continue
        row = dict(zip(cols, values))
        position_code = clean_text(row.get("position_code"))
        if not position_code:
            continue
        cleaned = {key: clean_text(value) for key, value in row.items()}
        cleaned["recruit_count"] = clean_int(row.get("recruit_count"))
        cleaned["raw_payload"] = row
        rows.append(cleaned)
    return rows


def match_position(major: Major, requirement: str | None) -> tuple[str, str, str] | None:
    req = compact_text(requirement)
    if not req:
        return None
    if any(token in req for token in ["不限", "无专业要求"]):
        return ("low", "unlimited_major", "岗位未限制具体专业，需核对其他资格条件")

    if len(major.normalized_name) >= 2 and major.normalized_name in req:
        return ("high", "exact_major", f"专业要求直接包含“{major.normalized_name}”")

    for category in major.categories:
        if compact_text(category) in req:
            return ("medium", "major_category", f"当前专业可归入“{category}”相关要求")

    for keyword in major.keywords:
        if len(keyword) >= 3 and keyword in req:
            return ("low", "keyword_related", f"专业关键词“{keyword}”出现在岗位专业要求中")

    return None


def build_matches(majors: list[Major], civil: list[dict[str, Any]], military: list[dict[str, Any]]) -> list[dict[str, Any]]:
    matches: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    positions = [
        ("civil_service", item["position_code"], item.get("major_requirement"))
        for item in civil
        if item.get("major_requirement")
    ] + [
        ("military_civilian", item["position_code"], item.get("major_requirement"))
        for item in military
        if item.get("major_requirement")
    ]
    for major in majors:
        for source, position_code, requirement in positions:
            match = match_position(major, requirement)
            if not match:
                continue
            level, match_type, reason = match
            key = (major.normalized_name, source, position_code, match_type)
            if key in matches:
                continue
            matches[key] = {
                "major_code": major.major_code,
                "major_name": major.major_name,
                "major_name_normalized": major.normalized_name,
                "position_source": source,
                "position_code": position_code,
                "match_level": level,
                "match_type": match_type,
                "match_reason": reason,
            }
    return list(matches.values())


def insert_sql(table: str, columns: list[str], rows: list[dict[str, Any]]) -> str:
    lines = [f"insert into public.{table} ({', '.join(columns)}) values"]
    values: list[str] = []
    for row in rows:
        vals = []
        for col in columns:
            value = row.get(col)
            vals.append(json_sql(value) if col == "raw_payload" else sql_literal(value))
        values.append("  (" + ", ".join(vals) + ")" + ("," if row is not rows[-1] else ""))
    lines.extend(values)
    if table == "civil_service_positions":
        lines.append("on conflict (position_code) do update set")
        updates = [f"  {col} = excluded.{col}" for col in columns if col != "position_code"]
        lines.append(",\n".join(updates) + ";")
    elif table == "military_civilian_positions":
        lines.append("on conflict (position_code) do update set")
        updates = [f"  {col} = excluded.{col}" for col in columns if col != "position_code"]
        lines.append(",\n".join(updates) + ";")
    else:
        lines.append("on conflict (major_name_normalized, position_source, position_code, match_type) do update set")
        lines.append("  major_code = excluded.major_code,\n  major_name = excluded.major_name,\n  match_level = excluded.match_level,\n  match_reason = excluded.match_reason;")
    return "\n".join(lines)


def write_chunks(civil: list[dict[str, Any]], military: list[dict[str, Any]], matches: list[dict[str, Any]]) -> None:
    CHUNK_DIR.mkdir(parents=True, exist_ok=True)
    for old in CHUNK_DIR.glob("*.sql"):
        old.unlink()

    (CHUNK_DIR / "001_prepare.sql").write_text(
        "\n".join([
            "delete from public.public_position_matches;",
            "delete from public.civil_service_positions;",
            "delete from public.military_civilian_positions;",
            "",
        ]),
        encoding="utf-8",
    )

    civil_cols = ["position_code", "exam_area", "department", "unit_name", "position_name", "recruit_count", "education_min", "degree_min", "major_requirement", "other_requirement", "essay_type", "raw_payload"]
    military_cols = ["position_code", "employer_no", "employer_name", "position_category", "position_name", "work_content", "recruit_count", "shortlist_ratio", "source_category", "education", "degree", "major_requirement", "exam_subject", "title_graduate", "title_social", "qualification_graduate", "qualification_social", "other_requirement", "work_location", "contact_phone", "raw_payload"]
    match_cols = ["major_code", "major_name", "major_name_normalized", "position_source", "position_code", "match_level", "match_type", "match_reason"]

    file_index = 2
    for start in range(0, len(civil), SQL_CHUNK_SIZE):
        chunk = civil[start:start + SQL_CHUNK_SIZE]
        (CHUNK_DIR / f"{file_index:03d}_civil_positions_{start + 1}_{start + len(chunk)}.sql").write_text(insert_sql("civil_service_positions", civil_cols, chunk), encoding="utf-8")
        file_index += 1
    for start in range(0, len(military), SQL_CHUNK_SIZE):
        chunk = military[start:start + SQL_CHUNK_SIZE]
        (CHUNK_DIR / f"{file_index:03d}_military_positions_{start + 1}_{start + len(chunk)}.sql").write_text(insert_sql("military_civilian_positions", military_cols, chunk), encoding="utf-8")
        file_index += 1
    for start in range(0, len(matches), MATCH_CHUNK_SIZE):
        chunk = matches[start:start + MATCH_CHUNK_SIZE]
        (CHUNK_DIR / f"{file_index:03d}_position_matches_{start + 1}_{start + len(chunk)}.sql").write_text(insert_sql("public_position_matches", match_cols, chunk), encoding="utf-8")
        file_index += 1

    (CHUNK_DIR / "999_summary.sql").write_text(
        "\n".join([
            "select",
            "  (select count(*) from public.civil_service_positions) as civil_positions,",
            "  (select count(*) from public.military_civilian_positions) as military_positions,",
            "  (select count(*) from public.public_position_matches) as position_matches;",
        ]),
        encoding="utf-8",
    )


def write_reports(civil: list[dict[str, Any]], military: list[dict[str, Any]], matches: list[dict[str, Any]], majors: list[Major]) -> None:
    level_counts = Counter(item["match_level"] for item in matches)
    source_counts = Counter(item["position_source"] for item in matches)
    major_counts = Counter(item["major_name"] for item in matches)
    summary = {
        "civil_positions": len(civil),
        "military_positions": len(military),
        "majors": len(majors),
        "matches": len(matches),
        "level_counts": dict(level_counts),
        "source_counts": dict(source_counts),
        "matched_major_count": len(major_counts),
        "top_matched_majors": major_counts.most_common(30),
    }
    SUMMARY_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    with MATCH_REVIEW_CSV.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow(["major_name", "match_count"])
        writer.writerows(major_counts.most_common())


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    majors = load_majors()
    civil = read_civil_positions()
    military = read_military_positions()
    matches = build_matches(majors, civil, military)
    write_chunks(civil, military, matches)
    write_reports(civil, military, matches, majors)
    print(SUMMARY_JSON.read_text(encoding="utf-8"))


if __name__ == "__main__":
    main()
